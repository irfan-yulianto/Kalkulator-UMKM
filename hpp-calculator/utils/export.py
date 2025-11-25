"""
Export utilities for Excel and PDF generation.
"""

import io
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel_report(
    calculation_result: Dict,
    product_name: str = "Produk",
    currency_symbol: str = "Rp"
) -> bytes:
    """
    Create Excel report with multiple sheets.

    Sheets:
    1. Summary - Overview of calculation results
    2. Ingredients - Detailed ingredient breakdown
    3. Cost_breakdown - Cost analysis

    Returns:
        Excel file as bytes
    """
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    title_font = Font(bold=True, size=14)
    currency_font = Font(size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== Sheet 1: Summary =====
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary['A1'] = f"Laporan HPP - {product_name}"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')

    # Date
    ws_summary['A2'] = f"Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_summary['A2'].font = Font(italic=True, size=10)

    # Get cost breakdown values (with defaults for backwards compatibility)
    material_cost = calculation_result.get('material_cost', calculation_result['total_batch_cost'])
    operational_cost = calculation_result.get('operational_cost', 0)
    other_cost = calculation_result.get('other_cost', 0)

    # Summary data
    summary_data = [
        ["", ""],
        ["RINGKASAN PERHITUNGAN", ""],
        ["Total Biaya per Batch", f"{currency_symbol} {calculation_result['total_batch_cost']:,.0f}".replace(",", ".")],
        ["  - Biaya Bahan Baku", f"{currency_symbol} {material_cost:,.0f}".replace(",", ".")],
        ["  - Biaya Operasional", f"{currency_symbol} {operational_cost:,.0f}".replace(",", ".")],
        ["  - Biaya Lain-lain", f"{currency_symbol} {other_cost:,.0f}".replace(",", ".")],
        ["Jumlah Output (porsi/unit)", calculation_result['output_units']],
        ["Target Margin", f"{calculation_result['target_margin_percent']:.1f}%"],
        ["", ""],
        ["HASIL PERHITUNGAN", ""],
        ["HPP per Porsi/Unit", f"{currency_symbol} {calculation_result['hpp_per_unit']:,.0f}".replace(",", ".")],
        ["Harga Jual Disarankan", f"{currency_symbol} {calculation_result['suggested_selling_price']:,.0f}".replace(",", ".")],
        ["", ""],
        ["ANALISIS MARGIN", ""],
        ["Harga Jual Aktual", f"{currency_symbol} {calculation_result['actual_selling_price']:,.0f}".replace(",", ".")],
        ["Margin Aktual", f"{calculation_result['actual_margin_percent']:.1f}%"],
        ["Gap vs Target", f"{calculation_result['gap_vs_target']:+.1f} pp"],
    ]

    for row_idx, row_data in enumerate(summary_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1 and value and value.isupper():
                cell.font = header_font
            elif col_idx == 1:
                cell.font = Font(size=11)
            else:
                cell.font = currency_font
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25

    # ===== Sheet 2: Bahan (Ingredients) =====
    ws_ingredients = wb.create_sheet("Bahan")

    # Headers - Indonesian names
    headers = ["Nama_Barang", "Qty_Total", "Satuan", "Harga_per_Unit", "Subtotal", "Kontribusi_%"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_ingredients.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, ing in enumerate(calculation_result['ingredients'], start=2):
        ws_ingredients.cell(row=row_idx, column=1, value=ing['name']).border = thin_border
        ws_ingredients.cell(row=row_idx, column=2, value=ing['quantity']).border = thin_border
        ws_ingredients.cell(row=row_idx, column=3, value=ing['unit']).border = thin_border
        ws_ingredients.cell(row=row_idx, column=4, value=ing['price_per_unit']).border = thin_border
        ws_ingredients.cell(row=row_idx, column=5, value=ing['line_cost']).border = thin_border
        ws_ingredients.cell(row=row_idx, column=6, value=ing['contribution_percent']).border = thin_border

    # Adjust column widths
    ws_ingredients.column_dimensions['A'].width = 20
    ws_ingredients.column_dimensions['B'].width = 12
    ws_ingredients.column_dimensions['C'].width = 10
    ws_ingredients.column_dimensions['D'].width = 15
    ws_ingredients.column_dimensions['E'].width = 15
    ws_ingredients.column_dimensions['F'].width = 12

    # ===== Sheet 3: Analisis Biaya (Cost Breakdown) =====
    ws_cost = wb.create_sheet("Analisis_Biaya")

    # Headers - Indonesian names (sorted by contribution)
    cost_headers = ["Nama_Barang", "Qty_Total", "Satuan", "Harga_per_Unit", "Subtotal", "Kontribusi_%"]
    for col_idx, header in enumerate(cost_headers, start=1):
        cell = ws_cost.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Sort by contribution (descending)
    sorted_ingredients = sorted(
        calculation_result['ingredients'],
        key=lambda x: x['contribution_percent'],
        reverse=True
    )

    for row_idx, ing in enumerate(sorted_ingredients, start=2):
        ws_cost.cell(row=row_idx, column=1, value=ing['name']).border = thin_border
        ws_cost.cell(row=row_idx, column=2, value=ing['quantity']).border = thin_border
        ws_cost.cell(row=row_idx, column=3, value=ing['unit']).border = thin_border
        ws_cost.cell(row=row_idx, column=4, value=ing['price_per_unit']).border = thin_border
        ws_cost.cell(row=row_idx, column=5, value=ing['line_cost']).border = thin_border
        ws_cost.cell(row=row_idx, column=6, value=ing['contribution_percent']).border = thin_border

    # Adjust column widths
    for col, width in [('A', 20), ('B', 12), ('C', 10), ('D', 15), ('E', 15), ('F', 12)]:
        ws_cost.column_dimensions[col].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_import_template() -> bytes:
    """
    Create Excel template for batch import.

    Columns:
    - Nama_Barang (text): Nama bahan
    - Qty_Bahan (number): Jumlah per kemasan (misal: 250 gram)
    - Satuan (text): Satuan bahan (gram, kg, ml, pcs, dll)
    - Qty_Jumlah (number): Jumlah kemasan yang dibeli
    - Harga (number): Harga per kemasan

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    example_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers - new structure
    headers = ["Nama_Barang", "Qty_Bahan", "Satuan", "Qty_Jumlah", "Harga"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Example data - Qty_Bahan (per kemasan), Satuan, Qty_Jumlah (jumlah kemasan), Harga (per kemasan)
    examples = [
        ["Tepung Terigu", 250, "gram", 2, 15000],       # 250g x 2 bungkus = Rp 30.000
        ["Ayam Karkas", 1, "kg", 3, 40000],             # 1kg x 3 = Rp 120.000
        ["Minyak Goreng", 1, "liter", 2, 20000],        # 1L x 2 = Rp 40.000
        ["Bumbu Marinasi", 100, "gram", 1, 25000],      # 100g x 1 = Rp 25.000
        ["Kemasan Box", 1, "pcs", 50, 1500],            # 1 pcs x 50 = Rp 75.000
        ["Gas LPG", 3, "kg", 1, 22000],                 # 3kg x 1 = Rp 22.000
    ]

    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = thin_border

    # Add empty rows for user input
    for row_idx in range(8, 20):
        for col_idx in range(1, 6):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    # Instructions sheet
    ws_info = wb.create_sheet("Petunjuk")
    instructions = [
        ["PETUNJUK PENGGUNAAN TEMPLATE"],
        [""],
        ["1. Isi data bahan pada sheet 'Template'"],
        ["2. Kolom yang wajib diisi:"],
        ["   - Nama_Barang: Nama bahan (max 100 karakter)"],
        ["   - Qty_Bahan: Jumlah per kemasan (misal: tepung 250 gram/bungkus)"],
        ["   - Satuan: Satuan bahan (gram, kg, ml, liter, pcs, dll)"],
        ["   - Qty_Jumlah: Jumlah kemasan yang dibeli (misal: beli 2 bungkus)"],
        ["   - Harga: Harga per kemasan (angka > 0)"],
        [""],
        ["RUMUS PERHITUNGAN:"],
        ["   Subtotal = Qty_Jumlah x Harga"],
        ["   (Qty_Bahan dan Satuan hanya sebagai referensi)"],
        [""],
        ["CONTOH:"],
        ["   Tepung 250 gram/bungkus, beli 2 bungkus @ Rp 15.000"],
        ["   Subtotal = 2 x 15.000 = Rp 30.000"],
        [""],
        ["3. Hapus contoh data (baris kuning) sebelum mengisi data Anda"],
        ["4. Simpan file dan upload ke aplikasi"],
        [""],
        ["SATUAN YANG DIDUKUNG:"],
        ["kg, gram, liter, ml, meter, cm, piece, pcs, pack, box, unit, porsi, buah, lembar, botol, kaleng, sachet, bungkus"],
    ]

    for row_idx, row_data in enumerate(instructions, start=1):
        cell = ws_info.cell(row=row_idx, column=1, value=row_data[0] if row_data else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=12)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws_info.column_dimensions['A'].width = 70

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_import_file(file_content: bytes, filename: str) -> Tuple[List[Dict], List[str]]:
    """
    Parse uploaded Excel/CSV file with new column structure.

    Expected columns:
    - Nama_Barang: Nama bahan
    - Qty_Bahan: Jumlah per kemasan (misal: 250 gram)
    - Satuan: Satuan bahan
    - Qty_Jumlah: Jumlah kemasan yang dibeli
    - Harga: Harga per kemasan

    Args:
        file_content: File content as bytes
        filename: Original filename

    Returns:
        Tuple of (ingredients_list, error_messages)
    """
    errors = []
    ingredients = []

    try:
        # Determine file type
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            df = pd.read_excel(io.BytesIO(file_content))

        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()

        # Map common column name variations - new structure
        column_mapping = {
            'nama_barang': ['nama_barang', 'ingredient', 'bahan', 'nama', 'nama_bahan', 'name'],
            'qty_bahan': ['qty_bahan', 'qty_per_batch', 'quantity', 'qty', 'kuantitas'],
            'satuan': ['satuan', 'unit', 'uom'],
            'qty_jumlah': ['qty_jumlah', 'jumlah', 'jml', 'amount', 'qty_amount'],
            'harga': ['harga', 'price', 'price_per_unit', 'harga_per_unit', 'harga_satuan']
        }

        # Find matching columns
        final_columns = {}
        for target, alternatives in column_mapping.items():
            for alt in alternatives:
                if alt in df.columns:
                    final_columns[target] = alt
                    break

        # Check required columns
        required = ['nama_barang', 'qty_bahan', 'satuan', 'qty_jumlah', 'harga']
        missing = [col for col in required if col not in final_columns]

        if missing:
            errors.append(f"Kolom tidak ditemukan: {', '.join(missing)}")
            return [], errors

        # Rename columns to standard names
        df = df.rename(columns={v: k for k, v in final_columns.items()})

        # Process each row
        for idx, row in df.iterrows():
            row_num = idx + 2  # Account for header row and 0-index

            ingredient_name = str(row.get('nama_barang', '')).strip()

            # Skip empty rows
            if not ingredient_name or ingredient_name.lower() == 'nan':
                continue

            # Validate and parse qty_bahan (quantity per package)
            try:
                qty_bahan = float(row.get('qty_bahan', 0))
                if qty_bahan <= 0:
                    errors.append(f"Baris {row_num}: Qty Bahan harus > 0")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Baris {row_num}: Qty Bahan tidak valid")
                continue

            # Validate unit
            satuan = str(row.get('satuan', '')).strip()
            if not satuan or satuan.lower() == 'nan':
                errors.append(f"Baris {row_num}: Satuan harus diisi")
                continue

            # Validate and parse qty_jumlah (number of packages)
            try:
                qty_jumlah = int(float(row.get('qty_jumlah', 0)))
                if qty_jumlah <= 0:
                    errors.append(f"Baris {row_num}: Qty Jumlah harus > 0")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Baris {row_num}: Qty Jumlah tidak valid")
                continue

            # Validate and parse harga (price per package)
            try:
                harga = float(row.get('harga', 0))
                if harga <= 0:
                    errors.append(f"Baris {row_num}: Harga harus > 0")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Baris {row_num}: Harga tidak valid")
                continue

            ingredients.append({
                'nama_barang': ingredient_name,
                'qty_bahan': qty_bahan,
                'satuan': satuan,
                'qty_jumlah': qty_jumlah,
                'harga': harga
            })

        if not ingredients and not errors:
            errors.append("Tidak ada data valid ditemukan dalam file")

    except Exception as e:
        errors.append(f"Error membaca file: {str(e)}")

    return ingredients, errors
